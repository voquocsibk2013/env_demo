"""
CO2 Footprint Calculator — Core Engine
Handles validation, lookup, calculation, and Excel export.
"""
from __future__ import annotations
import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# Schema definitions
# ──────────────────────────────────────────────────────────────────────────────

MTO_REQUIRED = {
    "Weight Item Descr.",
    "Material",
    "Cost Code COR",
    "Mod. Handl. Code",
    "Gross Dry Weight (kg)",
}

MEL_REQUIRED = {
    "Equipment type Description",
    "Cost Code COR",
    "Mod. Handl. Code",
    "Gross Dry Weight (kg)",
}

COR_REQUIRED = {"COR Code", "Category", "Description", "Emission factor"}


# ──────────────────────────────────────────────────────────────────────────────
# Data structures
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class ValidationError:
    file_name: str
    sheet_name: str
    row_number: Optional[int]
    column_name: Optional[str]
    problematic_value: Optional[str]
    explanation: str
    recommended_fix: str


@dataclass
class CalculationResult:
    success: bool
    status: str                        # 'success' | 'completed_with_errors' | 'failed'
    mto_total_tco2e: float = 0.0
    mel_total_tco2e: float = 0.0
    combined_total_tco2e: float = 0.0
    errors: list[ValidationError] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    detail_df: Optional[pd.DataFrame] = None
    mto_df: Optional[pd.DataFrame] = None
    mel_df: Optional[pd.DataFrame] = None
    cor_df: Optional[pd.DataFrame] = None


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _norm_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Trim & collapse whitespace in column headers."""
    df.columns = [" ".join(str(c).split()) for c in df.columns]
    return df


def _detect_sheet(
    all_sheets: dict[str, pd.DataFrame], required_cols: set[str]
) -> Optional[tuple[str, pd.DataFrame]]:
    """Return (name, df) for the first sheet whose headers satisfy required_cols."""
    for name, df in all_sheets.items():
        df = _norm_columns(df.copy())
        if required_cols.issubset(set(df.columns)):
            return name, df
    return None


def _is_excel_error(val) -> bool:
    error_strings = {"#ref!", "#value!", "#n/a", "#name?", "#div/0!", "#null!", "#num!"}
    return isinstance(val, str) and val.strip().lower() in error_strings


# ──────────────────────────────────────────────────────────────────────────────
# Loading
# ──────────────────────────────────────────────────────────────────────────────

def load_cor_lookup(source) -> tuple[Optional[pd.DataFrame], list[ValidationError]]:
    """
    Load COR code lookup from a file path, bytes/IO, or DataFrame.
    Returns (cor_df, errors).
    """
    errors: list[ValidationError] = []

    if isinstance(source, pd.DataFrame):
        df = _norm_columns(source.copy())
    else:
        try:
            raw = pd.read_excel(source, sheet_name=None)
        except Exception as e:
            errors.append(ValidationError(
                file_name=str(source), sheet_name="", row_number=None,
                column_name=None, problematic_value=None,
                explanation=f"Cannot open COR lookup file: {e}",
                recommended_fix="Ensure the COR code file is a valid .xlsx workbook."
            ))
            return None, errors

        result = _detect_sheet(raw, COR_REQUIRED)
        if result is None:
            errors.append(ValidationError(
                file_name=str(source), sheet_name="", row_number=None,
                column_name=None, problematic_value=None,
                explanation=f"COR lookup file has no sheet with required columns: {COR_REQUIRED}",
                recommended_fix="Ensure the COR lookup sheet contains: COR Code, Category, Description, Emission factor."
            ))
            return None, errors
        _, df = result

    # Drop rows where COR Code is NaN
    df = df.dropna(subset=["COR Code"])
    df["COR Code"] = df["COR Code"].astype(str).str.strip()

    # Validate emission factor is numeric
    df["Emission factor"] = pd.to_numeric(df["Emission factor"], errors="coerce")
    blank_ef = df[df["Emission factor"].isna()]
    for _, row in blank_ef.iterrows():
        errors.append(ValidationError(
            file_name="COR Lookup", sheet_name="COR_Code",
            row_number=None, column_name="Emission factor",
            problematic_value=str(row.get("COR Code")),
            explanation=f"COR code '{row['COR Code']}' has a blank or non-numeric Emission factor.",
            recommended_fix="Provide a numeric Emission factor for every COR code entry."
        ))

    return df, errors


def load_input_workbook(
    source,
    file_name: str = "upload.xlsx"
) -> tuple[Optional[pd.DataFrame], Optional[pd.DataFrame], list[ValidationError]]:
    """
    Load MTO and/or MEL sheets from an uploaded workbook.
    Returns (mto_df, mel_df, errors).
    """
    errors: list[ValidationError] = []

    try:
        all_sheets = pd.read_excel(source, sheet_name=None)
    except Exception as e:
        errors.append(ValidationError(
            file_name=file_name, sheet_name="", row_number=None,
            column_name=None, problematic_value=None,
            explanation=f"Cannot open workbook: {e}",
            recommended_fix="Ensure the uploaded file is a valid .xlsx workbook."
        ))
        return None, None, errors

    # Try to find MTO sheet
    mto_df = None
    result = _detect_sheet(all_sheets, MTO_REQUIRED)
    if result:
        sheet_name, mto_df = result
        mto_df = _norm_columns(mto_df.copy())
        mto_df["_source_sheet"] = sheet_name
        mto_df["_source_row"] = range(2, len(mto_df) + 2)  # 1-indexed, header=row1

    # Try to find MEL sheet
    mel_df = None
    result = _detect_sheet(all_sheets, MEL_REQUIRED)
    if result:
        sheet_name, mel_df = result
        mel_df = _norm_columns(mel_df.copy())
        mel_df["_source_sheet"] = sheet_name
        mel_df["_source_row"] = range(2, len(mel_df) + 2)

    if mto_df is None and mel_df is None:
        errors.append(ValidationError(
            file_name=file_name, sheet_name="", row_number=None,
            column_name=None, problematic_value=None,
            explanation="No valid MTO or MEL sheet found in the workbook.",
            recommended_fix=(
                f"Ensure at least one sheet contains MTO columns {MTO_REQUIRED} "
                f"or MEL columns {MEL_REQUIRED}."
            )
        ))

    return mto_df, mel_df, errors


# ──────────────────────────────────────────────────────────────────────────────
# Validation
# ──────────────────────────────────────────────────────────────────────────────

def _validate_rows(
    df: pd.DataFrame,
    cor_lookup: pd.DataFrame,
    file_name: str,
) -> tuple[pd.DataFrame, list[ValidationError]]:
    """
    Validate each row; add Status / Error_Message columns.
    Returns annotated df + list of errors.
    """
    errors: list[ValidationError] = []
    statuses = []
    messages = []

    cor_set = set(cor_lookup["COR Code"].astype(str).str.strip())

    for _, row in df.iterrows():
        sheet = row["_source_sheet"]
        src_row = int(row["_source_row"])
        row_errors = []

        # Weight
        weight_raw = row.get("Gross Dry Weight (kg)")
        if pd.isna(weight_raw):
            row_errors.append(
                f"Gross Dry Weight (kg) is blank."
            )
            errors.append(ValidationError(
                file_name=file_name, sheet_name=sheet, row_number=src_row,
                column_name="Gross Dry Weight (kg)", problematic_value=str(weight_raw),
                explanation=f"Sheet '{sheet}', row {src_row}: Gross Dry Weight (kg) is blank.",
                recommended_fix="Provide a numeric weight value."
            ))
        elif _is_excel_error(weight_raw):
            row_errors.append(f"Gross Dry Weight (kg) contains Excel error: {weight_raw}.")
            errors.append(ValidationError(
                file_name=file_name, sheet_name=sheet, row_number=src_row,
                column_name="Gross Dry Weight (kg)", problematic_value=str(weight_raw),
                explanation=f"Sheet '{sheet}', row {src_row}: Gross Dry Weight (kg) = '{weight_raw}' is an Excel error.",
                recommended_fix="Resolve the Excel formula error in the source file."
            ))
        else:
            try:
                float(weight_raw)
            except (ValueError, TypeError):
                row_errors.append(f"Gross Dry Weight (kg) = '{weight_raw}' is not numeric.")
                errors.append(ValidationError(
                    file_name=file_name, sheet_name=sheet, row_number=src_row,
                    column_name="Gross Dry Weight (kg)", problematic_value=str(weight_raw),
                    explanation=f"Sheet '{sheet}', row {src_row}: Gross Dry Weight (kg) = '{weight_raw}' is not numeric.",
                    recommended_fix="Replace with a valid numeric weight."
                ))

        # COR code
        cor_raw = row.get("Cost Code COR")
        if pd.isna(cor_raw) or str(cor_raw).strip() == "":
            row_errors.append("Cost Code COR is blank.")
            errors.append(ValidationError(
                file_name=file_name, sheet_name=sheet, row_number=src_row,
                column_name="Cost Code COR", problematic_value=str(cor_raw),
                explanation=f"Sheet '{sheet}', row {src_row}: Cost Code COR is blank.",
                recommended_fix="Provide a valid COR code."
            ))
        else:
            cor_val = str(cor_raw).strip()
            if cor_val not in cor_set:
                row_errors.append(f"COR code '{cor_val}' not found in lookup.")
                errors.append(ValidationError(
                    file_name=file_name, sheet_name=sheet, row_number=src_row,
                    column_name="Cost Code COR", problematic_value=cor_val,
                    explanation=f"Sheet '{sheet}', row {src_row}: COR code '{cor_val}' was not found in COR lookup.",
                    recommended_fix="Use a COR code that exists in the reference table."
                ))
            else:
                # Check emission factor
                ef_row = cor_lookup[cor_lookup["COR Code"] == cor_val]
                if not ef_row.empty and pd.isna(ef_row.iloc[0]["Emission factor"]):
                    row_errors.append(f"COR code '{cor_val}' has no Emission factor.")
                    errors.append(ValidationError(
                        file_name=file_name, sheet_name=sheet, row_number=src_row,
                        column_name="Emission factor", problematic_value=cor_val,
                        explanation=f"Sheet '{sheet}', row {src_row}: COR code '{cor_val}' exists but Emission factor is blank.",
                        recommended_fix="Add a numeric Emission factor for this COR code in the lookup."
                    ))

        # Mod. Handl. Code
        mhc_raw = row.get("Mod. Handl. Code")
        if pd.isna(mhc_raw) or str(mhc_raw).strip() == "":
            row_errors.append("Mod. Handl. Code is blank.")
            errors.append(ValidationError(
                file_name=file_name, sheet_name=sheet, row_number=src_row,
                column_name="Mod. Handl. Code", problematic_value=str(mhc_raw),
                explanation=f"Sheet '{sheet}', row {src_row}: Mod. Handl. Code is blank.",
                recommended_fix="Provide a valid Module Handling Code."
            ))

        if row_errors:
            statuses.append("ERROR")
            messages.append("; ".join(row_errors))
        else:
            statuses.append("VALID")
            messages.append("")

    df = df.copy()
    df["_status"] = statuses
    df["_error_msg"] = messages
    return df, errors


# ──────────────────────────────────────────────────────────────────────────────
# Calculation
# ──────────────────────────────────────────────────────────────────────────────

def _calculate(df: pd.DataFrame, cor_lookup: pd.DataFrame, data_type: str) -> pd.DataFrame:
    """
    For VALID rows: look up emission factor, compute tCO2e.
    Returns detail rows ready for output.
    """
    cor_map = cor_lookup.set_index("COR Code")[["Category", "Description", "Emission factor"]].to_dict("index")
    records = []

    for _, row in df.iterrows():
        cor_code = str(row.get("Cost Code COR", "")).strip()
        weight = row.get("Gross Dry Weight (kg)")
        status = row["_status"]
        error_msg = row["_error_msg"]

        category = description = emission_factor = emission_kgco2e = emission_tco2e = None

        if status == "VALID":
            lookup = cor_map.get(cor_code, {})
            category = lookup.get("Category")
            description = lookup.get("Description")
            emission_factor = lookup.get("Emission factor")
            if emission_factor is not None and not pd.isna(emission_factor):
                emission_kgco2e = float(weight) * float(emission_factor)
                emission_tco2e = emission_kgco2e / 1000

        item_desc = (
            row.get("Weight Item Descr.")
            if data_type == "MTO"
            else row.get("Equipment type Description")
        )
        material = row.get("Material") if data_type == "MTO" else None

        records.append({
            "Source": data_type,
            "Source Sheet": row["_source_sheet"],
            "Source Row": int(row["_source_row"]),
            "Item Description": item_desc,
            "Material": material,
            "COR Code": cor_code if cor_code else None,
            "Category": category,
            "COR Description": description,
            "Mod. Handl. Code": row.get("Mod. Handl. Code"),
            "Gross Dry Weight (kg)": weight,
            "Emission Factor": emission_factor,
            "Emission (kgCO2e)": emission_kgco2e,
            "Emission (tCO2e)": emission_tco2e,
            "Validation Status": status,
            "Error Message": error_msg if error_msg else None,
        })

    return pd.DataFrame(records)


# ──────────────────────────────────────────────────────────────────────────────
# Main entry point
# ──────────────────────────────────────────────────────────────────────────────

def calculate_footprint(
    input_source,
    cor_source,
    file_name: str = "upload.xlsx",
) -> CalculationResult:
    """
    Full pipeline: load → validate → calculate → return CalculationResult.

    Args:
        input_source : file path, bytes, or IO-like for the MTO/MEL workbook.
        cor_source   : file path, bytes, IO-like, or pd.DataFrame for COR lookup.
        file_name    : display name used in error messages.
    """
    result = CalculationResult(success=False, status="failed")

    # 1. Load COR lookup
    cor_df, cor_errors = load_cor_lookup(cor_source)
    result.errors.extend(cor_errors)
    if cor_df is None:
        result.status = "failed"
        return result
    result.cor_df = cor_df[["COR Code", "Category", "Description", "Emission factor"]].copy()

    # 2. Load input workbook
    mto_raw, mel_raw, load_errors = load_input_workbook(input_source, file_name)
    result.errors.extend(load_errors)
    if mto_raw is None and mel_raw is None:
        result.status = "failed"
        return result

    all_detail_frames = []
    critical_errors = False

    # 3. Validate & calculate MTO
    if mto_raw is not None:
        mto_validated, mto_errors = _validate_rows(mto_raw, cor_df, file_name)
        result.errors.extend(mto_errors)
        if mto_errors:
            critical_errors = True
        result.mto_df = _norm_columns(
            mto_raw.drop(columns=[c for c in mto_raw.columns if c.startswith("_")])
        )
        detail_mto = _calculate(mto_validated, cor_df, "MTO")
        result.mto_total_tco2e = detail_mto.loc[
            detail_mto["Validation Status"] == "VALID", "Emission (tCO2e)"
        ].sum()
        all_detail_frames.append(detail_mto)

    # 4. Validate & calculate MEL
    if mel_raw is not None:
        mel_validated, mel_errors = _validate_rows(mel_raw, cor_df, file_name)
        result.errors.extend(mel_errors)
        if mel_errors:
            critical_errors = True
        result.mel_df = _norm_columns(
            mel_raw.drop(columns=[c for c in mel_raw.columns if c.startswith("_")])
        )
        detail_mel = _calculate(mel_validated, cor_df, "MEL")
        result.mel_total_tco2e = detail_mel.loc[
            detail_mel["Validation Status"] == "VALID", "Emission (tCO2e)"
        ].sum()
        all_detail_frames.append(detail_mel)

    if all_detail_frames:
        result.detail_df = pd.concat(all_detail_frames, ignore_index=True)
        result.combined_total_tco2e = result.mto_total_tco2e + result.mel_total_tco2e

    if critical_errors:
        result.status = "completed_with_errors"
        result.success = True   # Partial success — workbook still saved
    else:
        result.status = "success"
        result.success = True

    return result


# ──────────────────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_ERROR_FILL   = PatternFill("solid", start_color="FFCCCC")
_VALID_FILL   = PatternFill("solid", start_color="C6EFCE")
_WARN_FILL    = PatternFill("solid", start_color="FFEB9C")
_BODY_FONT    = Font(name="Arial", size=10)
_THIN         = Side(style="thin", color="CCCCCC")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _write_df_to_sheet(ws, df: pd.DataFrame, highlight_status_col: Optional[str] = None):
    """Write a DataFrame to an openpyxl worksheet with formatting."""
    headers = list(df.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    status_col_idx = None
    if highlight_status_col and highlight_status_col in headers:
        status_col_idx = headers.index(highlight_status_col) + 1

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        status_val = None
        if status_col_idx:
            status_val = getattr(row, df.columns[status_col_idx - 1], None)

        row_fill = None
        if status_val == "ERROR":
            row_fill = _ERROR_FILL
        elif status_val == "VALID":
            row_fill = _VALID_FILL
        elif status_val == "WARNING":
            row_fill = _WARN_FILL

        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _BODY_FONT
            cell.border = _BORDER
            if row_fill:
                cell.fill = row_fill

    # Auto-width
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_workbook(result: CalculationResult, output_path) -> None:
    """
    Save the full calculation workbook to output_path (str or IO).
    Sheets: Input_MTO, Input_MEL, COR_Lookup_Used, Calculation_Detail,
            Summary, Validation_Log, Error_Report (if errors).
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Input_MTO ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Input_MTO")
    if result.mto_df is not None:
        _write_df_to_sheet(ws, result.mto_df)
    else:
        ws["A1"] = "No MTO data in upload."

    # ── Input_MEL ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Input_MEL")
    if result.mel_df is not None:
        _write_df_to_sheet(ws, result.mel_df)
    else:
        ws["A1"] = "No MEL data in upload."

    # ── COR_Lookup_Used ────────────────────────────────────────────────────
    ws = wb.create_sheet("COR_Lookup_Used")
    if result.cor_df is not None:
        _write_df_to_sheet(ws, result.cor_df)

    # ── Calculation_Detail ─────────────────────────────────────────────────
    ws = wb.create_sheet("Calculation_Detail")
    if result.detail_df is not None:
        _write_df_to_sheet(ws, result.detail_df, highlight_status_col="Validation Status")

    # ── Summary ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    _build_summary_sheet(ws, result)

    # ── Validation_Log ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Validation_Log")
    if result.warnings:
        for i, w in enumerate(result.warnings, 2):
            ws[f"A{i}"] = w
        ws["A1"] = "Warning"
        ws["A1"].font = _HEADER_FONT
        ws["A1"].fill = _HEADER_FILL
    else:
        ws["A1"] = "No warnings."

    # ── Error_Report ───────────────────────────────────────────────────────
    if result.errors:
        ws = wb.create_sheet("Error_Report")
        err_data = pd.DataFrame([
            {
                "File": e.file_name,
                "Sheet": e.sheet_name,
                "Row": e.row_number,
                "Column": e.column_name,
                "Problematic Value": e.problematic_value,
                "Explanation": e.explanation,
                "Recommended Fix": e.recommended_fix,
            }
            for e in result.errors
        ])
        _write_df_to_sheet(ws, err_data)

    wb.save(output_path)


def _build_summary_sheet(ws, result: CalculationResult):
    title_font = Font(bold=True, size=14, name="Arial", color="1F4E79")
    label_font = Font(bold=True, size=11, name="Arial")
    value_font = Font(size=11, name="Arial")
    section_fill = PatternFill("solid", start_color="D9E1F2")

    ws["A1"] = "CO2 Footprint — Calculation Summary"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Run status: {result.status.upper()}"
    ws["A2"].font = Font(bold=True, size=11, name="Arial",
                         color="00B050" if result.status == "success" else "FF0000")

    rows = [
        ("", ""),
        ("Source", "Total Footprint (tCO2e)"),
        ("MTO", result.mto_total_tco2e),
        ("MEL", result.mel_total_tco2e),
        ("Combined Total", result.combined_total_tco2e),
    ]
    for r_offset, (label, val) in enumerate(rows, 4):
        a = ws.cell(row=r_offset, column=1, value=label)
        b = ws.cell(row=r_offset, column=2, value=val)
        a.font = label_font
        b.font = value_font
        if label in ("Source", "Combined Total"):
            a.fill = section_fill
            b.fill = section_fill
        if isinstance(val, float):
            b.number_format = '#,##0.000'

    # Subtotals by Category
    if result.detail_df is not None and not result.detail_df.empty:
        valid_rows = result.detail_df[result.detail_df["Validation Status"] == "VALID"]
        if not valid_rows.empty:
            row_start = 12
            ws.cell(row=row_start, column=1, value="Subtotals by Category").font = label_font
            ws.cell(row=row_start, column=1).fill = section_fill
            ws.cell(row=row_start, column=2, value="tCO2e").font = label_font
            ws.cell(row=row_start, column=2).fill = section_fill
            ws.cell(row=row_start, column=3, value="Source").font = label_font
            ws.cell(row=row_start, column=3).fill = section_fill

            by_cat = (
                valid_rows.groupby(["Category", "Source"])["Emission (tCO2e)"]
                .sum().reset_index()
            )
            for i, row in by_cat.iterrows():
                r = row_start + 1 + i
                ws.cell(row=r, column=1, value=row["Category"]).font = value_font
                ws.cell(row=r, column=2, value=row["Emission (tCO2e)"]).number_format = '#,##0.000'
                ws.cell(row=r, column=3, value=row["Source"]).font = value_font

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
